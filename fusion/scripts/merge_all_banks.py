
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook

ROOT = Path.home() / "DTCup_Prep"
EXPORT_DIR = ROOT / "fusion" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

SOURCE_REPO_NAMES = [
    "dtcup-2026-prep",
]
REPOS = [ROOT / name for name in SOURCE_REPO_NAMES if (ROOT / name).exists()]
ENABLE_DEDUP = False

TEXT_EXTS = {".md", ".txt", ".tex"}
JSON_EXTS = {".json"}
XLSX_EXTS = {".xlsx"}
GENERATED_BANK_NAME = "DTCUP Mega Bank"
DTCUP_2026_PREP_MAIN_BANK = Path("真题题库/真题-按知识点分类.md")
DTCUP_2026_PREP_MULTI_BANK = Path("真题题库/真题-易错多选题.md")
DTCUP_2026_PREP_REPEAT_BANK = Path("真题题库/真题-高频重复题.md")
QBANK_XLSX = Path("2025dtCup-Qbank/output.xlsx")

OPTION_RE = re.compile(r"^\s*([A-D])[\.\、:：\)\s]+(.+?)\s*$", re.I)
MARKDOWN_OPTION_RE = re.compile(r"^\s*-\s*\*\*([A-D])\.\*\*\s*(.+?)(?:\s*✓)?\s*$", re.I)
ANSWER_RE = re.compile(r"(?:[【\[]?\s*(?:答案|answer|正确答案)\s*[】\]]?)\s*[:：]?\s*(.+?)\s*$", re.I)
QSTART_RE = re.compile(r"^\s*(\d{1,4})[\.\、\)]\s*(.+)$")
MARKDOWN_QSTART_RE = re.compile(r"^\s*(?:\*\*\[[^]]+\]\*\*\s*)?(\d{1,4})[\.\、\)]\s*(.+)$")
SOURCE_MARKDOWN_QSTART_RE = re.compile(r"^\s*\*\*\[来源：([^]]+)\]\*\*\s*(\d{1,4})[\.\、\)]\s*(.+)$")
REPEAT_SOURCE_RE = re.compile(r"^###\s*出现试卷[:：]\s*(.+)$")
LATEX_CHOICE_RE = re.compile(r"\\begin\{choice\}\{([^}]*)\}\s*\[\s*\](.*?)\\end\{choice\}", re.DOTALL)
LATEX_TASKS_RE = re.compile(r"\\begin\{tasks\}\([^)]*\)(.*?)\\end\{tasks\}", re.DOTALL)
LATEX_TASK_RE = re.compile(r"\\task\s*([^\n\\]*(?:\n(?![\\])[^\n\\]*)*)")
QUESTION_NUMBER_PREFIX_RE = re.compile(r"^\s*\d{1,4}[\.\、\)]\s*")

def add_question(bank: List[Dict[str, Any]], question: str, options: List[str], answer: Any, explanation: str = ""):
    question = question.strip()
    options = [str(x).strip() for x in options if str(x).strip()]
    if not question or len(options) < 2:
        return
    qtype = "single"
    if len(options) == 2 and set(options) == {"正确", "错误"}:
        qtype = "judge"
    if isinstance(answer, list) and len(answer) > 1:
        qtype = "multiple"
    if answer is None:
        answer = [0] if qtype == "multiple" else 0
    bank.append({
        "type": qtype,
        "question": question,
        "options": options,
        "answer": answer,
        "explanation": explanation,
    })

def parse_answer_token(token: str):
    token = str(token).strip()
    token = token.replace("**", "").replace("✓", "").replace("，", "").replace(",", "")
    token = re.sub(r"\s+", "", token)
    if not token:
        return None
    if token in ("正确", "对"):
        return 0
    if token in ("错误", "错"):
        return 1
    if re.fullmatch(r"[A-Da-d]", token):
        return ord(token.upper()) - ord("A")
    if re.fullmatch(r"[A-Da-d]{2,4}", token):
        return [ord(c.upper()) - ord("A") for c in token]
    if token.isdigit():
        n = int(token)
        return max(0, n - 1)
    return None


def source_label(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def find_header(header: Dict[str, int], *aliases: str) -> Optional[int]:
    for alias in aliases:
        col = header.get(normalize_header(alias))
        if col:
            return col
    return None


def clean_latex_text(text: str) -> str:
    cleaned = str(text or "")
    cleaned = re.sub(r"%.*", "", cleaned)
    cleaned = re.sub(r"\\href\{[^}]*\}\{([^}]*)\}", r"\1", cleaned)
    cleaned = re.sub(r"\\textbf\{([^}]*)\}", r"\1", cleaned)
    cleaned = re.sub(r"\\[a-zA-Z]+\*?(?:\[[^\]]*\])?\{([^{}]*)\}", r"\1", cleaned)
    cleaned = re.sub(r"\\[a-zA-Z]+\*?", " ", cleaned)
    cleaned = cleaned.replace(r"\_", "_")
    cleaned = cleaned.replace("{", " ").replace("}", " ")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip()


def normalize_text(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip())


def normalize_question_for_match(text: Any) -> str:
    normalized = normalize_text(text)
    return QUESTION_NUMBER_PREFIX_RE.sub("", normalized)


def normalize_question_for_lookup(text: Any) -> str:
    normalized = normalize_question_for_match(text)
    normalized = normalized.replace("（", "(").replace("）", ")")
    normalized = re.sub(r"\(\s*\)|（\s*）", "", normalized)
    normalized = re.sub(r"[，。！？、；：,.!?;:（）()\[\]\s]+", "", normalized)
    return normalized.lower()


def question_signature(question: Dict[str, Any]) -> tuple[Any, ...]:
    answer = question.get("answer")
    if isinstance(answer, list):
        answer_key: Any = tuple(sorted(answer))
    else:
        answer_key = answer
    return (
        normalize_text(question.get("question", "")),
        tuple(normalize_text(option) for option in question.get("options", [])),
        answer_key,
    )


def question_match_signature(question: Dict[str, Any]) -> tuple[Any, ...]:
    answer = question.get("answer")
    if isinstance(answer, list):
        answer_key: Any = tuple(sorted(answer))
    else:
        answer_key = answer
    return (
        normalize_question_for_match(question.get("question", "")),
        tuple(normalize_text(option) for option in question.get("options", [])),
        answer_key,
    )


def deduplicate_questions(bank: List[Dict[str, Any]]) -> tuple[List[Dict[str, Any]], int]:
    seen = set()
    deduped = []
    removed = 0
    for item in bank:
        signature = question_signature(item)
        if signature in seen:
            removed += 1
            continue
        seen.add(signature)
        deduped.append(item)
    return deduped, removed

def import_json(path: Path) -> List[Dict[str, Any]]:
    out = []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except:
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig"))
        except:
            return out

    def walk(obj):
        if isinstance(obj, dict):
            if "question" in obj and "options" in obj:
                add_question(
                    out,
                    str(obj.get("question", "")),
                    list(obj.get("options", [])),
                    obj.get("answer"),
                    f"来源: {source_label(path)}",
                )
            else:
                for v in obj.values():
                    walk(v)
        elif isinstance(obj, list):
            for x in obj:
                walk(x)

    walk(data)
    return out

def import_xlsx(path: Path) -> List[Dict[str, Any]]:
    out = []
    try:
        wb = load_workbook(path, data_only=True)
    except:
        return out

    for ws in wb.worksheets:
        header = {}
        start_row = None
        for r in range(1, min(6, ws.max_row) + 1):
            vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, min(ws.max_column, 12) + 1)]
            if any(x in " ".join(vals) for x in ["题目", "答案", "A", "B", "question", "answer"]):
                for c in range(1, ws.max_column + 1):
                    header[normalize_header(ws.cell(r, c).value or "")] = c
                start_row = r + 1
                break
        if start_row is None:
            continue

        cq = find_header(header, "题目", "question", "问题")
        ca = find_header(header, "答案", "answer")
        cA = find_header(header, "A", "A选项")
        cB = find_header(header, "B", "B选项")
        cC = find_header(header, "C", "C选项")
        cD = find_header(header, "D", "D选项")

        if not cq or not cA or not cB:
            continue

        for r in range(start_row, ws.max_row + 1):
            q = str(ws.cell(r, cq).value or "").strip()
            if not q:
                continue
            options = []
            for c in [cA, cB, cC, cD]:
                if c:
                    v = str(ws.cell(r, c).value or "").strip()
                    if v:
                        options.append(v)
            ans = parse_answer_token(ws.cell(r, ca).value) if ca else None
            add_question(out, q, options, ans, f"来源: {source_label(path)} | sheet={ws.title}")
    return out

def read_text(path: Path) -> str:
    for enc in ("utf-8", "utf-8-sig", "gb18030", "gbk", "latin1"):
        try:
            return path.read_text(encoding=enc)
        except:
            pass
    return ""


def is_generated_bank_json(path: Path) -> bool:
    if path.suffix.lower() != ".json":
        return False
    text = read_text(path)
    if not text:
        return False
    try:
        data = json.loads(text)
    except:
        return False
    return isinstance(data, dict) and data.get("name") == GENERATED_BANK_NAME


def import_latex_choice_text(path: Path, text: str) -> List[Dict[str, Any]]:
    out = []
    for match in LATEX_CHOICE_RE.finditer(text):
        answer = parse_answer_token(match.group(1))
        body = match.group(2)
        tasks_match = LATEX_TASKS_RE.search(body)
        if not tasks_match:
            continue
        stem = clean_latex_text(body[:tasks_match.start()] + " " + body[tasks_match.end():])
        options = []
        for task in LATEX_TASK_RE.finditer(tasks_match.group(1)):
            option = clean_latex_text(task.group(1))
            if option:
                options.append(option)
        add_question(out, stem, options, answer, f"来源: {source_label(path)}")
    return out


def clean_markdown_text(text: str) -> str:
    cleaned = str(text or "").strip()
    cleaned = cleaned.replace("**", "")
    return normalize_text(cleaned)


def merge_explanation(existing: str, extra: str) -> str:
    parts = []
    for text in (existing, extra):
        cleaned = str(text or "").strip()
        if cleaned and cleaned not in parts:
            parts.append(cleaned)
    return "\n".join(parts)


def answer_label(question: Dict[str, Any]) -> str:
    labels = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    answer = question.get("answer")
    options = question.get("options", [])
    if question.get("type") == "judge":
        index = int(answer) if isinstance(answer, int) else 0
        return options[index] if 0 <= index < len(options) else "正确"
    if isinstance(answer, list):
        parts = []
        for idx in sorted(answer):
            if 0 <= idx < len(options):
                parts.append(f"{labels[idx]}. {options[idx]}")
        return "；".join(parts)
    if isinstance(answer, int) and 0 <= answer < len(options):
        return f"{labels[answer]}. {options[answer]}"
    return ""


def answer_value_text(question: Dict[str, Any]) -> str:
    answer = question.get("answer")
    options = question.get("options", [])
    if question.get("type") == "judge":
        index = int(answer) if isinstance(answer, int) else 0
        return options[index] if 0 <= index < len(options) else "正确"
    if isinstance(answer, list):
        values = [options[idx] for idx in sorted(answer) if 0 <= idx < len(options)]
        return " ".join(values)
    if isinstance(answer, int) and 0 <= answer < len(options):
        return options[answer]
    return ""


def load_qbank_lookup(path: Path) -> Dict[str, Dict[str, str]]:
    if not path.exists():
        return {}
    lookup: Dict[str, Dict[str, str]] = {}
    try:
        wb = load_workbook(path, data_only=True)
    except:
        return lookup

    for ws in wb.worksheets:
        header = {
            normalize_header(ws.cell(1, c).value or ""): c
            for c in range(1, ws.max_column + 1)
        }
        cq = find_header(header, "问题", "题目", "question")
        c_key = find_header(header, "考点")
        c_lookup = find_header(header, "答案速查")
        if not cq:
            continue
        for r in range(2, ws.max_row + 1):
            question = str(ws.cell(r, cq).value or "").strip()
            if not question:
                continue
            key = normalize_question_for_lookup(question)
            if not key:
                continue
            lookup[key] = {
                "keypoint": str(ws.cell(r, c_key).value or "").strip() if c_key else "",
                "answer_lookup": str(ws.cell(r, c_lookup).value or "").strip() if c_lookup else "",
            }
    return lookup


def load_note_snippets(note_paths: List[Path]) -> List[str]:
    snippets: List[str] = []
    for path in note_paths:
        text = read_text(path)
        if not text:
            continue
        for raw_line in text.splitlines():
            line = clean_markdown_text(raw_line)
            if not line or line.startswith("#"):
                continue
            line = re.sub(r"^[>\-\*\d\.\s]+", "", line)
            line = clean_markdown_text(line)
            if len(line) < 8:
                continue
            snippets.append(line)
    return snippets


def extract_query_terms(question: Dict[str, Any], keypoint: str) -> List[str]:
    source = " ".join(
        [
            normalize_question_for_match(question.get("question", "")),
            keypoint,
            answer_label(question),
            answer_value_text(question),
        ]
    )
    source = re.sub(r"([A-Za-z0-9])([\u4e00-\u9fff])", r"\1 \2", source)
    source = re.sub(r"([\u4e00-\u9fff])([A-Za-z0-9])", r"\1 \2", source)
    tokens = re.findall(r"[\u4e00-\u9fffA-Za-z0-9\-\+]+", source)
    stopwords = {"以下", "哪项", "哪一项", "下列", "属于", "不属于", "的是", "正确", "错误", "功能", "系统"}
    terms: List[str] = []

    def add_term(term: str) -> None:
        term = term.strip()
        if len(term) < 2 or term in stopwords or term in terms:
            return
        terms.append(term)

    for token in tokens:
        add_term(token)
        parts = re.findall(r"[A-Za-z0-9\-\+]+|[\u4e00-\u9fff]+", token)
        for part in parts:
            add_term(part)
            if re.fullmatch(r"[\u4e00-\u9fff]{2,}", part):
                max_window = min(4, len(part))
                for size in range(2, max_window + 1):
                    for start in range(0, len(part) - size + 1):
                        add_term(part[start : start + size])
    terms.sort(key=len, reverse=True)
    return terms[:16]


def find_best_note_snippet(question: Dict[str, Any], keypoint: str, snippets: List[str]) -> str:
    terms = extract_query_terms(question, keypoint)
    answer_text = answer_value_text(question)
    best_snippet = ""
    best_score = 0
    best_knowledge_snippet = ""
    best_knowledge_score = 0
    question_markers = (
        "真题原题",
        "真题多选",
        "真题判断",
        "反复出现的真题",
        "跨届反复出现的真题",
        "今日真题自测",
        "→",
        "？",
        "?",
    )
    knowledge_markers = (
        "负责",
        "用于",
        "包括",
        "不包含",
        "包含",
        "获取",
        "管理",
        "建立",
        "修改",
        "删除",
        "由",
        "决定",
        "范围",
        "周期广播",
        "逻辑隔离",
        "物理隔离",
        "通过",
        "连接",
        "接口",
        "作用",
        "功能",
        "触发",
        "配置",
        "属于",
        "选择",
    )
    for snippet in snippets:
        score = 0
        is_questionish = any(marker in snippet for marker in question_markers)
        for term in terms:
            if term in snippet:
                score += max(2, len(term))
        if keypoint and keypoint in snippet:
            score += 6
        if question.get("type") != "judge" and answer_text and answer_text in snippet:
            score += len(answer_text) + 4
        if any(marker in snippet for marker in knowledge_markers):
            score += 5
        if question.get("type") == "judge":
            if answer_text == "错误" and any(marker in snippet for marker in ("不", "无", "不是", "错误")):
                score += 4
            if answer_text == "正确" and any(marker in snippet for marker in ("可以", "支持", "可采用", "竞争", "非竞争")):
                score += 2
        if is_questionish:
            score -= 8
        if re.match(r"^\d+[\.\、\)]", snippet):
            score -= 4
        if len(snippet) <= 48:
            score += 1
        if score > best_score:
            best_score = score
            best_snippet = snippet
        if not is_questionish and score > best_knowledge_score:
            best_knowledge_score = score
            best_knowledge_snippet = snippet
    if best_knowledge_score >= 16:
        return best_knowledge_snippet
    if best_knowledge_score >= max(10, int(best_score * 0.5)):
        return best_knowledge_snippet
    return best_snippet if best_score >= 2 else ""


def render_note_reason(note_snippet: str, question: Dict[str, Any]) -> str:
    snippet = normalize_text(note_snippet)
    snippet = re.sub(
        r"^\s*(?:真题原题|真题判断|真题多选|反复出现的真题|跨届反复出现的真题)[:：]\s*",
        "",
        snippet,
    )
    snippet = re.sub(r"\s*\[[^\]]+\]\s*$", "", snippet)
    snippet = re.sub(r"[（(]考了\d+次\+?[)）]", "", snippet)
    answer_text = answer_value_text(question)
    qtype = question.get("type")

    if "|" in snippet:
        cells = [normalize_text(cell) for cell in snippet.split("|") if normalize_text(cell)]
        if len(cells) == 2:
            left, right = cells
            if re.fullmatch(r"[A-Za-z0-9\-\+]+", left) and re.search(r"[\u4e00-\u9fff]", right):
                return f"{left}负责{right}"
            if re.fullmatch(r"[A-Za-z0-9\-\+]+", right) and re.search(r"[\u4e00-\u9fff]", left):
                return f"{right}负责{left}"
            if answer_text and answer_text in left and right:
                return f"{right}对应{left}"
            if answer_text and answer_text in right and left:
                return f"{left}对应{right}"
            return "、".join(cells)
        if len(cells) >= 3:
            subject_tokens = re.findall(r"\b([A-Za-z]{2,}\d*)\b", normalize_question_for_match(question.get("question", "")))
            subject = max(subject_tokens, key=len, default="")
            if subject and cells[0] in {"主要内容", "功能", "作用"}:
                return f"{subject}的{cells[0]}包括{cells[1]}"
        snippet = "、".join(cells)

    if "→" in snippet:
        left, _right = [normalize_text(part) for part in snippet.split("→", 1)]
        if qtype == "judge":
            return left.rstrip("。；，？！?")
        if answer_text:
            if "通知UE获取" in left:
                left = re.sub(r"(通知UE获取)(?:哪个消息|什么|哪些消息)?", rf"\1{answer_text}", left)
                return left.rstrip("。；，？！?")
            function_match = re.search(r"(.+?)的功能", left)
            if function_match:
                return f"{function_match.group(1)}的功能是{answer_text}"
            return f"{answer_text}是该题考查的核心结论"

    return snippet.rstrip("。；，")


def has_any(text: str, keywords: List[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def option_texts(question: Dict[str, Any]) -> List[str]:
    return [normalize_text(option) for option in question.get("options", [])]


def build_manual_priority_explanation(question: Dict[str, Any], metadata: Dict[str, str]) -> str:
    question_text = normalize_question_for_match(question.get("question", ""))
    question_lower = question_text.lower()
    answer_text = answer_label(question)
    answer_value = answer_value_text(question)
    answer_lower = answer_value.lower()
    keypoint = metadata.get("keypoint", "").strip()
    options = option_texts(question)
    options_join = " ".join(options)
    options_lower = options_join.lower()
    parts: List[str] = []
    if keypoint:
        parts.append(f"考点：{keypoint}。")

    if "pc5" in question_lower and has_any(options_join, ["基站", "UE之间", "PC5", "Uu"]):
        parts.append(
            f"因为PC5是UE与UE直接通信的接口，覆盖V2V/V2I/V2P，不经过基站；车与基站走Uu口，车与网络通信也走Uu口，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "c-v2x" in question_lower and has_any(question_text, ["两种通信接口", "通信接口"]) and has_any(
        options_lower, ["pc5", "uu"]
    ):
        parts.append(f"因为C-V2X提供两类接口：PC5用于直连通信，Uu用于经基站和核心网通信，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if "mec" in question_lower and "rru" in options_lower:
        parts.append(
            f"因为MEC可以与基站共址，也可部署在接入汇聚机房或骨干汇聚机房，但不能与RRU共址，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "v2x" in question_lower and has_any(question_text, ["不属于V2X范畴", "不属于V2X"]) and "v2c" in options_lower:
        parts.append(
            f"因为V2X常见形态是V2V、V2I、V2P、V2N，分别对应车与车、基础设施、行人、网络；V2C不在常规定义内，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "pc5" in question_lower and "sdap" in options_lower:
        parts.append(
            f"因为SDAP是5G NR新增的协议层，基于LTE实现的PC5协议栈里没有SDAP，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "pdcch" in question_lower and has_any(options_join, ["QPSK", "16QAM", "64QAM", "256QAM"]):
        parts.append(
            f"因为PDCCH是控制信道，采用Polar码编码，调制方式固定为QPSK；PDSCH作为数据信道才支持更高阶QAM，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "pdsch" in question_lower and has_any(options_join, ["256QAM", "64QAM", "512QAM", "128QAM", "LDPC", "Polar"]):
        parts.append(
            f"因为PDSCH是数据信道，R15最高支持256QAM并采用LDPC编码；PDCCH/PBCH这类控制信道通常走QPSK和Polar，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "cce" in question_lower and "reg" in question_lower:
        parts.append(
            f"因为控制信道资源关系是 1CCE = 6REG，RE和RB都不是CCE的直接等价单位，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "pci" in question_lower:
        sss_match = re.search(r"SSS\s*ID\s*为?\s*(\d+)", question_text, re.I)
        pss_match = re.search(r"PSS\s*ID\s*为?\s*(\d+)", question_text, re.I)
        if sss_match and pss_match:
            nid1 = int(sss_match.group(1))
            nid2 = int(pss_match.group(1))
            pci = 3 * nid1 + nid2
            parts.append(
                f"因为PCI = 3 × N_ID_1 + N_ID_2，其中N_ID_1由SSS确定、N_ID_2由PSS确定，所以 3 × {nid1} + {nid2} = {pci}，所以选 {answer_text}。"
            )
            return "核心解析： " + " ".join(parts)
        if has_any(question_text, ["范围", "取值"]) and has_any(options_join, ["0-1007", "0-503", "1-1008"]):
            parts.append(
                f"因为PCI = 3 × N_ID_1 + N_ID_2，N_ID_1范围是0到335，N_ID_2范围是0到2，所以PCI总共有1008个取值，范围是0到1007，所以选 {answer_text}。"
            )
            return "核心解析： " + " ".join(parts)

    if has_any(question_lower, ["pss", "sss"]) and has_any(question_text, ["小区搜索", "下行同步"]):
        parts.append(f"因为UE完成小区搜索和下行同步主要依赖PSS和SSS；PBCH/MIB是在同步后再解调获取，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if "uu" in question_lower and has_any(question_text, ["控制面协议", "用户面协议", "次序"]):
        if "控制面" in question_text:
            parts.append(
                f"因为Uu口控制面从上到下是RRC-PDCP-RLC-MAC-PHY；SDAP属于用户面，SCTP/IP不属于空口无线协议栈，所以选 {answer_text}。"
            )
        else:
            parts.append(
                f"因为Uu口用户面从上到下是SDAP-PDCP-RLC-MAC-PHY；RRC属于控制面，SCTP/IP不属于空口无线协议栈，所以选 {answer_text}。"
            )
        return "核心解析： " + " ".join(parts)

    if has_any(question_text, ["回传"]) and has_any(options_join, ["CU-5GC", "DU-CU", "AAU-DU"]):
        parts.append(
            f"因为三段传输里前传一般是AAU-DU，中传一般是DU-CU，回传才是CU-5GC，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if has_any(question_text, ["核心网通过哪个接口相连", "基站和核心网通过哪个接口"]) and has_any(
        options_join, ["NG", "Xn", "S1", "X2"]
    ):
        parts.append(
            f"因为gNB与5GC之间通过NG接口相连；Xn是基站与基站之间的接口，S1/X2属于LTE体系，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "xn" in question_lower and has_any(question_text, ["第一条信令"]) and has_any(
        options_join, ["Handover Request", "Handover Required"]
    ):
        if "向目标基站发送" in question_text:
            parts.append(
                f"因为Xn切换时源基站发给目标基站的第一条信令是Handover Request；Handover Required是面向AMF的NG切换信令，所以选 {answer_text}。"
            )
        else:
            parts.append(
                f"因为涉及核心网侧时，源基站发往AMF的第一条信令是Handover Required；Handover Request是发给目标基站的，所以选 {answer_text}。"
            )
        return "核心解析： " + " ".join(parts)

    if "随机森林" in question_text:
        parts.append(
            f"因为随机森林本质上是多棵决策树做并行集成，采用Bagging思路训练；Boosting强调串行纠错，不是随机森林核心机制，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "bagging" in question_lower and "pasting" in question_lower:
        parts.append(
            f"因为Bagging是有放回采样，同一样本可以被重复抽到；Pasting是不放回采样，所以两者区别就在采样是否放回，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "aau" in question_lower and has_any(question_text, ["单模集束光纤", "路由长度"]) and "100" in options_lower:
        parts.append(
            f"因为单模集束光纤在BBU与AAU路由长度超过100米时使用，距离更短时通常不需要切到这类方案，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "aau" in question_lower and has_any(question_text, ["功耗", "整机"]) and "90" in options_lower:
        parts.append(f"因为AAU功耗约占整机90%，是基站节能优化的主要对象，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if has_any(question_lower, ["rrc状态", "rrc状态转换"]) and has_any(
        options_join, ["RRC空闲态", "RRC非活动态", "RRC连接"]
    ):
        parts.append(
            f"因为RRC状态只有IDLE、INACTIVE、CONNECTED三种，IDLE不能直接进入INACTIVE；通常是IDLE先建链到CONNECTED，再进入INACTIVE，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "rrc重建" in question_lower and has_any(question_text, ["上下文不能被恢复", "不能被恢复但有资源"]):
        if "setup" in answer_lower:
            parts.append(
                f"因为UE上下文不能恢复但还有资源时，基站会转入重新建立连接流程，向UE发送RRC Setup；若直接拒绝才会回RRC Reject，所以选 {answer_text}。"
            )
        else:
            parts.append(
                f"因为RRC重建时若基站直接拒绝UE重建和新建，才会回RRC Reject，不会进入重新建立连接流程，所以选 {answer_text}。"
            )
        return "核心解析： " + " ".join(parts)

    if "rrc inact" in question_lower and has_any(question_text, ["恢复业务", "触发哪条信令"]):
        parts.append(
            f"因为RRC INACTIVE会保留上下文，恢复业务时应发送RRC resume request，而不是重新完整发起RRC建立，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "srb" in question_lower and has_any(options_join, ["SRB1", "SRB2", "SRB3", "SRB0"]):
        if "仅承载" in question_text and "NAS" in question_text:
            parts.append(
                f"因为SRB2只承载NAS消息并映射到DCCH；SRB1既承载RRC也承载NAS，SRB0映射到CCCH，所以选 {answer_text}。"
            )
            return "核心解析： " + " ".join(parts)
        if has_any(question_text, ["不仅承载NAS消息还可以承载RRC消息", "承载RRC消息"]):
            parts.append(
                f"因为SRB1映射到DCCH，既可以承载RRC消息也可以承载NAS消息；SRB2只承载NAS，所以选 {answer_text}。"
            )
            return "核心解析： " + " ".join(parts)

    if "srb3" in question_lower and has_any(question_text, ["未建立", "测量结果", "上报给网络"]):
        parts.append(
            f"因为SRB3未建立时，SCG测量结果要通过UL Information Transfer MRDC上报；不是直接走Measurement Report或重配置完成消息，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "prach" in question_lower and has_any(question_text, ["format0-3", "zc序列长度"]) and "839" in options_lower:
        parts.append(
            f"因为PRACH format0-3对应的ZC序列长度是839；139对应短序列，不是format0-3，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "prach" in question_lower and has_any(question_text, ["长prach格式", "支持几种长prach格式"]) and "4" in options_lower:
        parts.append(
            f"因为5G NR长PRACH格式共有4种，短PRACH格式是另一套分类，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "AMF" in question_text and has_any(answer_value, ["注册管理", "接入管理", "移动性管理"]):
        details = ["AMF负责注册管理、接入管理和移动性管理"]
        if any("会话" in option for option in options):
            details.append("SMF负责会话建立、修改、删除")
        if any("下行数据" in option for option in options):
            details.append("SMF还负责下行数据通知，所以下行数据通知也不属于AMF")
        if any("IP" in option for option in options):
            details.append("SMF还负责UE IP分配，所以UE IP分配不属于AMF")
        parts.append(f"因为{'；'.join(details)}，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if "SMF" in question_text and has_any(answer_value, ["会话", "IP", "下行数据"]):
        parts.append(
            f"因为SMF负责会话建立、修改、删除、UE IP分配和下行数据通知，而注册管理、接入管理、移动性管理属于AMF，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "mib" in question_lower and has_any(question_text, ["内容", "包含"]) and has_any(
        options_join, ["小区ID", "PHICH", "CORESET", "SFN"]
    ):
        parts.append(
            f"因为MIB包含SFN和SIB1的CORESET配置，不包含小区ID和PHICH；小区ID由PSS/SSS获取，PHICH是LTE概念，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "mib" in question_lower and has_any(question_text, ["作用", "获取"]) and "sib1" in options_lower:
        parts.append(
            f"因为MIB最重要的作用是通知UE如何获取SIB1，而RACH、TDD等详细系统信息由SIB1继续提供，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if "nsa" in question_lower and has_any(options_lower, ["option3x", "option 3x", "option2", "option 2"]):
        parts.append(
            f"因为我国商用NSA采用Option 3x，Option 2属于SA独立组网并使用5GC，不是NSA商用部署方式，所以选 {answer_text}。"
        )
        return "核心解析： " + " ".join(parts)

    if has_any(question_text, ["子帧", "时隙数"]) and "3" in options:
        parts.append(f"因为每子帧时隙数等于 2^μ，只可能是1、2、4、8、16，3不在合法取值内，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if "60khz" in answer_lower and has_any(question_text, ["常规CP", "扩展CP", "同时支持"]):
        parts.append(f"因为只有60kHz同时支持Normal CP和Extended CP，其它常见SCS不同时满足这两个条件，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if "ssb" in question_lower and has_any(options_join, ["CRS", "PSS", "SSS", "PBCH"]):
        parts.append(f"因为SSB只包括PSS、SSS、PBCH，CRS是LTE参考信号，不属于5G SSB，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if has_any(question_text, ["网络切片", "切片"]) and has_any(" ".join(options), ["逻辑", "物理"]):
        parts.append(f"因为网络切片是在同一硬件基础设施上划分多个虚拟端到端网络，切片之间是逻辑隔离，不是物理隔离，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if has_any(question_text, ["软硬件解耦", "软件和硬件解耦"]) and has_any(options_join, ["NFV", "SDN", "MEC"]):
        parts.append(f"因为NFV负责网络功能虚拟化，实现软硬件解耦；SDN强调控制与转发分离，MEC强调边缘计算，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if has_any(question_text, ["随机接入"]) and has_any(question_text, ["切换", "波束恢复", "初始接入", "链路失败"]):
        if question.get("type") == "judge":
            verdict = "正确" if answer_value == "正确" else "错误"
            relation = "一致" if verdict == "正确" else "不一致"
            parts.append(
                f"因为初始接入、无线链路失败后初始接入通常走竞争随机接入，而切换和波束恢复可以采用非竞争随机接入，所以题干表述与规则{relation}，判断为{verdict}。"
            )
        else:
            parts.append(
                f"因为初始接入、无线链路失败后初始接入通常走竞争随机接入，而切换和波束恢复可以采用非竞争随机接入，所以选 {answer_text}。"
            )
        return "核心解析： " + " ".join(parts)

    if has_any(question_lower, ["rrc inactive", "inactive恢复", "恢复业务"]) and "resume" in answer_lower:
        parts.append(f"因为RRC INACTIVE保留上下文，恢复业务时走RRC resume request，不需要重新完整建链，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    if has_any(question_lower, ["option 8", "cu", "du"]) and has_any(question_text, ["时延", "带宽"]):
        parts.append(f"因为Option 8分割点最靠近PHY，对前传时延最严格、传输带宽要求也最高，所以选 {answer_text}。")
        return "核心解析： " + " ".join(parts)

    return ""


def build_core_explanation(question: Dict[str, Any], metadata: Dict[str, str], snippets: List[str]) -> str:
    keypoint = metadata.get("keypoint", "").strip()
    answer_text = answer_label(question)
    answer_value = answer_value_text(question)
    manual_explanation = build_manual_priority_explanation(question, metadata)
    if manual_explanation:
        return manual_explanation

    parts: List[str] = []
    if keypoint:
        parts.append(f"考点：{keypoint}。")
    note_snippet = find_best_note_snippet(question, keypoint, snippets)
    if note_snippet:
        note_reason = render_note_reason(note_snippet, question)
        if question.get("type") == "judge":
            verdict = "正确" if answer_value == "正确" else "错误"
            relation = "一致" if verdict == "正确" else "不一致"
            parts.append(f"因为{note_reason}，与题干表述{relation}，所以判断为{verdict}。")
        elif question.get("type") == "multiple":
            parts.append(f"因为{note_reason}，能直接对应正确选项 {answer_text}，所以应选 {answer_text}。")
        else:
            parts.append(f"因为{note_reason}，能直接对应正确项 {answer_text}，所以选 {answer_text}。")
    elif metadata.get("answer_lookup"):
        if question.get("type") == "judge":
            parts.append(f"因为该考点结论与题干核对后应判为{answer_value}。")
        elif question.get("type") == "multiple":
            parts.append(f"因为题目考查的核心结论对应 {metadata['answer_lookup']}，所以应选 {answer_text}。")
        else:
            parts.append(f"因为题目考查的核心结论对应 {metadata['answer_lookup']}，所以选 {answer_text}。")
    elif answer_text:
        if question.get("type") == "judge":
            parts.append(f"因为题干与该知识点结论核对后应判为{answer_value}。")
        else:
            parts.append(f"因为只有 {answer_text} 符合该知识点结论，所以选 {answer_text}。")
    return "核心解析： " + " ".join(parts) if parts else ""


def enrich_missing_explanations(
    bank: List[Dict[str, Any]],
    qbank_path: Optional[Path] = None,
    note_paths: Optional[List[Path]] = None,
) -> None:
    qbank_lookup = load_qbank_lookup(qbank_path or (ROOT / QBANK_XLSX))
    if note_paths is None:
        note_paths = sorted((ROOT / "dtcup-2026-prep").glob("Phase*/*.md"))
    note_snippets = load_note_snippets(note_paths)

    for item in bank:
        explanation = str(item.get("explanation", "")).strip()
        if "辨析：" in explanation or "核心解析：" in explanation:
            continue

        metadata = qbank_lookup.get(normalize_question_for_lookup(item.get("question", "")), {})
        core_explanation = build_core_explanation(item, metadata, note_snippets)
        if core_explanation:
            item["explanation"] = merge_explanation(explanation, core_explanation)


def import_generated_markdown_questions(path: Path, include_source: bool) -> List[Dict[str, Any]]:
    text = read_text(path)
    if not text:
        return []

    lines = [x.rstrip() for x in text.splitlines()]
    out = []
    current_q = ""
    current_options: Dict[str, str] = {}
    current_answer = None
    current_source: Optional[str] = None
    current_repeat_context: Optional[str] = None
    current_explanation_lines: List[str] = []
    pending_repeat_context: Optional[str] = None

    def flush():
        nonlocal current_q, current_options, current_answer, current_source, current_repeat_context, current_explanation_lines
        if current_q:
            if current_options:
                options = [current_options[k] for k in ["A", "B", "C", "D"] if k in current_options]
            elif current_answer in (0, 1):
                options = ["正确", "错误"]
            else:
                options = []

            if options:
                explanation = ""
                if include_source and current_source:
                    explanation = f"来源: {current_source}"
                if current_repeat_context:
                    explanation = merge_explanation(explanation, f"高频重复题: 出现试卷：{current_repeat_context}")
                if current_explanation_lines:
                    explanation = merge_explanation(explanation, "\n".join(current_explanation_lines))
                add_question(out, current_q, options, current_answer, explanation)

        current_q = ""
        current_options = {}
        current_answer = None
        current_source = None
        current_repeat_context = None
        current_explanation_lines = []

    for line in lines:
        s = line.strip()
        if not s:
            continue

        repeat_match = REPEAT_SOURCE_RE.match(s)
        if repeat_match:
            flush()
            pending_repeat_context = repeat_match.group(1).strip()
            continue

        source_qm = SOURCE_MARKDOWN_QSTART_RE.match(s)
        if source_qm:
            flush()
            current_source = source_qm.group(1).strip()
            current_q = f"{source_qm.group(2)}. {source_qm.group(3)}".strip()
            pending_repeat_context = None
            continue

        qm = QSTART_RE.match(s) or MARKDOWN_QSTART_RE.match(s)
        if qm:
            flush()
            current_q = f"{qm.group(1)}. {qm.group(2)}".strip()
            pending_repeat_context = None
            continue

        if pending_repeat_context and not current_q:
            current_q = clean_markdown_text(s)
            current_repeat_context = pending_repeat_context
            pending_repeat_context = None
            continue

        om = OPTION_RE.match(s) or MARKDOWN_OPTION_RE.match(s)
        if om and current_q:
            current_options[om.group(1).upper()] = om.group(2).strip()
            continue

        am = ANSWER_RE.search(s)
        if am and current_q:
            current_answer = parse_answer_token(am.group(1))
            continue

        if current_q and current_answer is not None:
            current_explanation_lines.append(clean_markdown_text(s))
            continue

        if current_q and not current_options:
            current_q += " " + clean_markdown_text(s)

    flush()
    return out


def merge_question_explanations(bank: List[Dict[str, Any]], enrichments: List[Dict[str, Any]]) -> None:
    by_signature: Dict[tuple[Any, ...], List[str]] = {}
    for item in enrichments:
        explanation = str(item.get("explanation", "")).strip()
        if not explanation:
            continue
        by_signature.setdefault(question_match_signature(item), []).append(explanation)

    for item in bank:
        for explanation in by_signature.get(question_match_signature(item), []):
            item["explanation"] = merge_explanation(item.get("explanation", ""), explanation)


def import_dtcup_2026_prep_bank(repo: Path) -> List[Dict[str, Any]]:
    main_questions = import_generated_markdown_questions(repo / DTCUP_2026_PREP_MAIN_BANK, include_source=True)
    multi_explanations = import_generated_markdown_questions(repo / DTCUP_2026_PREP_MULTI_BANK, include_source=False)
    repeat_notes = import_generated_markdown_questions(repo / DTCUP_2026_PREP_REPEAT_BANK, include_source=False)

    merge_question_explanations(main_questions, multi_explanations)
    merge_question_explanations(main_questions, repeat_notes)
    enrich_missing_explanations(main_questions)
    return main_questions


def import_line_text(path: Path, text: str) -> List[Dict[str, Any]]:
    lines = [x.rstrip() for x in text.splitlines()]

    out = []
    current_q = ""
    current_options = {}
    current_answer = None

    def flush():
        nonlocal current_q, current_options, current_answer
        if current_q and current_options:
            options = [current_options[k] for k in ["A", "B", "C", "D"] if k in current_options]
            add_question(out, current_q, options, current_answer, f"来源: {source_label(path)}")
        current_q = ""
        current_options = {}
        current_answer = None

    for line in lines:
        s = line.strip()
        if not s:
            continue

        qm = QSTART_RE.match(s) or MARKDOWN_QSTART_RE.match(s)
        if qm:
            flush()
            current_q = f"{qm.group(1)}. {qm.group(2)}".strip()
            continue

        om = OPTION_RE.match(s) or MARKDOWN_OPTION_RE.match(s)
        if om and current_q:
            current_options[om.group(1).upper()] = om.group(2).strip()
            continue

        am = ANSWER_RE.search(s)
        if am and current_q:
            current_answer = parse_answer_token(am.group(1))
            continue

        if current_q and not current_options:
            current_q += " " + s

    flush()
    return out

def import_text(path: Path) -> List[Dict[str, Any]]:
    text = read_text(path)
    if not text:
        return []
    if path.suffix.lower() == ".tex" and r"\begin{choice}" in text:
        items = import_latex_choice_text(path, text)
        if items:
            return items
    return import_line_text(path, text)

def main():
    bank = []
    stats = {}
    for repo in REPOS:
        count_before = len(bank)
        if repo.name == "dtcup-2026-prep":
            bank.extend(import_dtcup_2026_prep_bank(repo))
            stats[repo.name] = len(bank) - count_before
            continue
        for path in repo.rglob("*"):
            if not path.is_file():
                continue
            if is_generated_bank_json(path):
                continue
            suffix = path.suffix.lower()
            if suffix in JSON_EXTS:
                bank.extend(import_json(path))
            elif suffix in XLSX_EXTS:
                bank.extend(import_xlsx(path))
            elif suffix in TEXT_EXTS:
                bank.extend(import_text(path))
        stats[repo.name] = len(bank) - count_before

    removed_duplicates = 0
    if ENABLE_DEDUP:
        bank, removed_duplicates = deduplicate_questions(bank)

    out = {"name": "DTCUP Mega Bank", "questions": bank}
    out_json = EXPORT_DIR / "mega_bank.json"
    out_report = EXPORT_DIR / "mega_bank_report.txt"
    out_json.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    with out_report.open("w", encoding="utf-8") as f:
        for k, v in stats.items():
            f.write(f"{k}: {v}\n")
        f.write(f"REMOVED_DUPLICATES: {removed_duplicates}\n")
        f.write(f"TOTAL: {len(bank)}\n")
        f.write(f"OUTPUT: {out_json}\n")

    print(f"Done. Total questions: {len(bank)}")
    print(f"JSON: {out_json}")
    print(f"REPORT: {out_report}")

if __name__ == "__main__":
    main()
