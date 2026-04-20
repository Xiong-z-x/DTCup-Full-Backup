from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from fusion.scripts import merge_all_banks


def test_main_imports_all_supported_text_json_and_xlsx_files(tmp_path: Path) -> None:
    root = tmp_path / "DTCup_Prep"
    repo = root / "sample-repo"
    export_dir = root / "fusion" / "exports"
    repo.mkdir(parents=True)
    export_dir.mkdir(parents=True)

    (repo / "notes.md").write_text(
        "\n".join(
            [
                "1. 这是一道 Markdown 题目",
                "A. 选项一",
                "B. 选项二",
                "答案: A",
            ]
        ),
        encoding="utf-8",
    )

    (repo / "bank.json").write_text(
        json.dumps(
            {
                "items": [
                    {
                        "question": "这是一道 JSON 题目",
                        "options": ["选项一", "选项二"],
                        "answer": 0,
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["题目", "A", "B", "答案"])
    worksheet.append(["这是一道 XLSX 题目", "选项一", "选项二", "A"])
    workbook.save(repo / "sheet.xlsx")

    merge_all_banks.ROOT = root
    merge_all_banks.EXPORT_DIR = export_dir
    merge_all_banks.REPOS = [repo]

    merge_all_banks.main()

    mega_bank = json.loads((export_dir / "mega_bank.json").read_text(encoding="utf-8"))
    questions = {item["question"] for item in mega_bank["questions"]}

    assert questions == {
        "1. 这是一道 Markdown 题目",
        "这是一道 JSON 题目",
        "这是一道 XLSX 题目",
    }


def test_import_text_parses_markdown_bank_export(tmp_path: Path) -> None:
    path = tmp_path / "bank.md"
    path.write_text(
        "\n".join(
            [
                "# 大唐杯真题",
                "",
                "**[来源：示例试卷]** 1. Markdown 导出的题目",
                "",
                "- **A.** 选项一",
                "- **B.** 选项二 ✓",
                "- **C.** 选项三",
                "- **D.** 选项四",
                "【答案】B",
            ]
        ),
        encoding="utf-8",
    )

    items = merge_all_banks.import_text(path)

    assert len(items) == 1
    assert items[0]["question"] == "1. Markdown 导出的题目"
    assert items[0]["options"] == ["选项一", "选项二", "选项三", "选项四"]
    assert items[0]["answer"] == 1


def test_import_generated_markdown_questions_preserves_judge_and_explanation(tmp_path: Path) -> None:
    path = tmp_path / "bank.md"
    path.write_text(
        "\n".join(
            [
                "# 大唐杯真题",
                "",
                "**[来源：第十届示例试卷]** 1. 判断题题干",
                "",
                "【答案】✓ 正确",
                "",
                "**辨析：** 这是一段解析。",
            ]
        ),
        encoding="utf-8",
    )

    items = merge_all_banks.import_generated_markdown_questions(path, include_source=True)

    assert len(items) == 1
    assert items[0]["question"] == "1. 判断题题干"
    assert items[0]["type"] == "judge"
    assert items[0]["options"] == ["正确", "错误"]
    assert items[0]["answer"] == 0
    assert "来源: 第十届示例试卷" in items[0]["explanation"]
    assert "辨析： 这是一段解析。" in items[0]["explanation"]


def test_import_text_parses_latex_choice_tasks_format(tmp_path: Path) -> None:
    path = tmp_path / "bank.tex"
    path.write_text(
        "\n".join(
            [
                r"\begin{choice}{C}[]",
                "5G NR 结构中，1 个子帧的时隙数不可能有",
                r"\begin{tasks}(4)",
                r"\task 2",
                r"\task 1",
                r"\task 3",
                r"\task 4",
                r"\end{tasks}",
                r"\end{choice}",
            ]
        ),
        encoding="utf-8",
    )

    items = merge_all_banks.import_text(path)

    assert len(items) == 1
    assert items[0]["question"] == "5G NR 结构中，1 个子帧的时隙数不可能有"
    assert items[0]["options"] == ["2", "1", "3", "4"]
    assert items[0]["answer"] == 2


def test_import_xlsx_supports_problem_style_headers(tmp_path: Path) -> None:
    path = tmp_path / "bank.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["考点", "题型", "问题", "A选项", "B选项", "C选项", "D选项", "答案"])
    worksheet.append(["知识点", "单选", "这是一道表格题", "选项一", "选项二", "选项三", "选项四", "D"])
    workbook.save(path)

    items = merge_all_banks.import_xlsx(path)

    assert len(items) == 1
    assert items[0]["question"] == "这是一道表格题"
    assert items[0]["options"] == ["选项一", "选项二", "选项三", "选项四"]
    assert items[0]["answer"] == 3


def test_import_dtcup_2026_prep_bank_merges_explanations_without_adding_duplicates(tmp_path: Path) -> None:
    repo = tmp_path / "dtcup-2026-prep"
    bank_dir = repo / "真题题库"
    bank_dir.mkdir(parents=True)

    (bank_dir / "真题-按知识点分类.md").write_text(
        "\n".join(
            [
                "# 分类题库",
                "",
                "**[来源：卷一]** 1. 多选题干",
                "",
                "- **A.** 甲 ✓",
                "- **B.** 乙",
                "- **C.** 丙 ✓",
                "- **D.** 丁",
                "【答案】AC",
                "",
                "**[来源：卷二]** 2. 判断题干",
                "",
                "【答案】错误",
            ]
        ),
        encoding="utf-8",
    )
    (bank_dir / "真题-易错多选题.md").write_text(
        "\n".join(
            [
                "# 易错多选",
                "",
                "**[来源：卷一]** 1. 多选题干",
                "",
                "- **A.** 甲 ✓",
                "- **B.** 乙",
                "- **C.** 丙 ✓",
                "- **D.** 丁",
                "【答案】AC",
                "",
                "**辨析：** 这是多选解析。",
            ]
        ),
        encoding="utf-8",
    )
    (bank_dir / "真题-高频重复题.md").write_text(
        "\n".join(
            [
                "# 高频重复",
                "",
                "### 出现试卷：卷二；卷三",
                "",
                "判断题干",
                "",
                "【答案】错误",
            ]
        ),
        encoding="utf-8",
    )

    items = merge_all_banks.import_dtcup_2026_prep_bank(repo)

    assert len(items) == 2

    multiple = next(item for item in items if item["question"] == "1. 多选题干")
    judge = next(item for item in items if item["question"] == "2. 判断题干")

    assert multiple["answer"] == [0, 2]
    assert "来源: 卷一" in multiple["explanation"]
    assert "辨析： 这是多选解析。" in multiple["explanation"]

    assert judge["type"] == "judge"
    assert judge["answer"] == 1
    assert "来源: 卷二" in judge["explanation"]
    assert "高频重复题: 出现试卷：卷二；卷三" in judge["explanation"]


def test_enrich_missing_explanations_uses_qbank_keypoint_and_note_snippet(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["考点", "题型", "问题", "A选项", "B选项", "C选项", "D选项", "答案", "答案速查"])
    worksheet.append(
        [
            "网络架构演进",
            "单选",
            "在5G网络架构中，以下选项哪一项是AMF的功能？（  ）",
            "注册管理",
            "漫游功能",
            "会话建立修改删除",
            "下行数据通知",
            "A",
            "注册管理",
        ]
    )
    qbank_path = tmp_path / "output.xlsx"
    workbook.save(qbank_path)

    note_path = tmp_path / "Day2.md"
    note_path.write_text(
        "\n".join(
            [
                "# Day 2",
                "- AMF负责注册管理、接入管理和移动性管理。",
                "- SMF负责会话建立、修改、删除。",
            ]
        ),
        encoding="utf-8",
    )

    bank = [
        {
            "type": "single",
            "question": "1. 在5G网络架构中，以下选项哪一项是AMF的功能？",
            "options": ["注册管理", "漫游功能", "会话建立修改删除", "下行数据通知"],
            "answer": 0,
            "explanation": "来源: 卷一",
        }
    ]

    merge_all_banks.enrich_missing_explanations(bank, qbank_path=qbank_path, note_paths=[note_path])

    explanation = bank[0]["explanation"]
    assert "考点：网络架构演进" in explanation
    assert "因为" in explanation
    assert "AMF负责注册管理" in explanation
    assert "所以选 A. 注册管理" in explanation


def test_build_core_explanation_for_judge_uses_consistent_reasoning() -> None:
    question = {
        "type": "judge",
        "question": "1. MIB包含小区ID。",
        "options": ["正确", "错误"],
        "answer": 1,
    }
    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "5G协议与信令", "answer_lookup": ""},
        ["MIB包含SFN和SIB1的CORESET配置，不包含小区ID，小区ID由PSS和SSS获取。"],
    )

    assert "因为" in explanation
    assert "不包含小区ID" in explanation
    assert "所以判断为错误" in explanation


def test_build_core_explanation_prefers_knowledge_snippet_over_question_replay() -> None:
    question = {
        "type": "single",
        "question": "1. 在5G网络架构中，以下选项哪一项是AMF的功能？",
        "options": ["注册管理", "漫游功能", "会话建立修改删除", "下行数据通知"],
        "answer": 0,
    }
    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "网络架构演进", "answer_lookup": "注册管理"},
        [
            "真题原题：在5G网络架构中，以下哪项是AMF的功能？→ A. 注册管理 [第九届A组]",
            "AMF=注册管理，SMF=会话管理",
        ],
    )

    assert "AMF负责注册管理" in explanation
    assert "真题原题" not in explanation
    assert "所以选 A. 注册管理" in explanation


def test_build_core_explanation_uses_contrastive_reasoning_for_amf() -> None:
    question = {
        "type": "single",
        "question": "1. 在5G网络架构中，以下选项哪一项是AMF的功能？",
        "options": ["注册管理", "漫游功能", "会话建立修改删除", "下行数据通知"],
        "answer": 0,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "网络架构演进", "answer_lookup": "注册管理"},
        [],
    )

    assert "AMF负责注册管理" in explanation
    assert "SMF负责会话建立、修改、删除" in explanation
    assert "所以下行数据通知也不属于AMF" in explanation
    assert "所以选 A. 注册管理" in explanation


def test_build_core_explanation_uses_contrastive_reasoning_for_mib_content() -> None:
    question = {
        "type": "single",
        "question": "1. 关于MIB消息内容，下列哪项正确？",
        "options": ["包含小区ID", "包含PHICH配置", "包含SIB1的CORESET配置", "包含RACH详细配置"],
        "answer": 2,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "系统消息", "answer_lookup": "SIB1的CORESET配置"},
        [],
    )

    assert "MIB包含SFN和SIB1的CORESET配置" in explanation
    assert "小区ID由PSS/SSS获取" in explanation
    assert "PHICH是LTE概念" in explanation
    assert "所以选 C. 包含SIB1的CORESET配置" in explanation


def test_build_core_explanation_uses_contrastive_reasoning_for_nsa_option() -> None:
    question = {
        "type": "single",
        "question": "1. 目前我国5G商用网采用的NSA组网部署方式为？",
        "options": ["Option 3x", "Option 2", "Option 1", "Option 8"],
        "answer": 0,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "SA/NSA组网", "answer_lookup": "Option 3x"},
        [],
    )

    assert "我国商用NSA采用Option 3x" in explanation
    assert "Option 2属于SA" in explanation
    assert "所以选 A. Option 3x" in explanation


def test_build_core_explanation_uses_contrastive_reasoning_for_pc5() -> None:
    question = {
        "type": "single",
        "question": "1. C-V2X中PC5接口表述错误的是？",
        "options": ["是UE与基站之间的参考点", "UE之间的参考点", "包括基于LTE的PC5", "包括基于NR的PC5"],
        "answer": 0,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "V2X车联网", "answer_lookup": "UE之间"},
        [],
    )

    assert "PC5是UE与UE直接通信" in explanation
    assert "车与基站走Uu口" in explanation
    assert "所以选 A. 是UE与基站之间的参考点" in explanation


def test_build_core_explanation_uses_contrastive_reasoning_for_pdcch() -> None:
    question = {
        "type": "single",
        "question": "1. 5G NR系统中，PDCCH信道采用Polar码信道编码方式，调制方式为？",
        "options": ["16QAM", "256QAM", "QPSK", "64QAM"],
        "answer": 2,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "NR原理与关键技术", "answer_lookup": "QPSK"},
        [],
    )

    assert "PDCCH是控制信道" in explanation
    assert "Polar码" in explanation
    assert "PDSCH作为数据信道" in explanation
    assert "所以选 C. QPSK" in explanation


def test_build_core_explanation_uses_formula_reasoning_for_pci() -> None:
    question = {
        "type": "single",
        "question": "1. 5G NR中，SSS ID为102，PSS ID为2，此时PCI为？",
        "options": ["302", "308", "1008", "504"],
        "answer": 1,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "5G无线网络参数规划", "answer_lookup": "308"},
        [],
    )

    assert "PCI = 3 × N_ID_1 + N_ID_2" in explanation
    assert "3 × 102 + 2 = 308" in explanation
    assert "所以选 B. 308" in explanation


def test_build_core_explanation_uses_contrastive_reasoning_for_backhaul() -> None:
    question = {
        "type": "single",
        "question": "1. 5G网络架构中，CU/DU分离场景下，回传一般指哪两个之间的数据传送？",
        "options": ["AAU-DU", "AAU-BBU", "CU-5GC", "DU-CU"],
        "answer": 2,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "网络架构演进", "answer_lookup": "CU-5GC"},
        [],
    )

    assert "前传一般是AAU-DU" in explanation
    assert "中传一般是DU-CU" in explanation
    assert "回传才是CU-5GC" in explanation
    assert "所以选 C. CU-5GC" in explanation


def test_build_core_explanation_uses_contrastive_reasoning_for_bagging_vs_pasting() -> None:
    question = {
        "type": "single",
        "question": "1. bagging和pasting的区别是？",
        "options": ["预测器不同", "bagging采样不放回，pasting采样放回", "二者无区别", "bagging采样放回，pasting采样不放回"],
        "answer": 3,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "AI与机器学习", "answer_lookup": "bagging采样放回，pasting采样不放回"},
        [],
    )

    assert "Bagging是有放回采样" in explanation
    assert "Pasting是不放回采样" in explanation
    assert "所以选 D. bagging采样放回，pasting采样不放回" in explanation


def test_build_core_explanation_uses_reasoning_for_aau_fiber_length() -> None:
    question = {
        "type": "single",
        "question": "1. 对于大唐5G设备，单模集束光纤在BBU与AAU路由长度超过多少米使用？",
        "options": ["40", "60", "80", "100"],
        "answer": 3,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "基站SA开通与调测", "answer_lookup": "100"},
        [],
    )

    assert "单模集束光纤在BBU与AAU路由长度超过100米时使用" in explanation
    assert "所以选 D. 100" in explanation


def test_build_core_explanation_uses_reasoning_for_v2x_scope() -> None:
    question = {
        "type": "single",
        "question": "1. C-V2X中，下面不属于V2X范畴的是？",
        "options": ["V2V", "V2I", "V2N", "V2C"],
        "answer": 3,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "V2X车联网", "answer_lookup": "V2C"},
        [],
    )

    assert "V2X常见形态是V2V、V2I、V2P、V2N" in explanation
    assert "V2C不在常规定义内" in explanation
    assert "所以选 D. V2C" in explanation


def test_build_core_explanation_uses_reasoning_for_srb2() -> None:
    question = {
        "type": "single",
        "question": "1. 5G NR中哪类SRB仅承载NAS消息，映射到DCCH信道？",
        "options": ["SRB2", "SRB0", "SRB3", "SRB1"],
        "answer": 0,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "5G协议与信令", "answer_lookup": "SRB2"},
        [],
    )

    assert "SRB2只承载NAS消息" in explanation
    assert "SRB1既承载RRC也承载NAS" in explanation
    assert "所以选 A. SRB2" in explanation


def test_build_core_explanation_uses_reasoning_for_prach_zc_length() -> None:
    question = {
        "type": "single",
        "question": "1. 在5G NR PRACH规划中，PRACH格式format0-3对应ZC序列长度为？",
        "options": ["838", "139", "138", "839"],
        "answer": 3,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "NR原理与关键技术", "answer_lookup": "839"},
        [],
    )

    assert "PRACH format0-3对应的ZC序列长度是839" in explanation
    assert "139对应短序列" in explanation
    assert "所以选 D. 839" in explanation


def test_build_core_explanation_uses_reasoning_for_rrc_rebuild_setup() -> None:
    question = {
        "type": "single",
        "question": "1. 5G NR系统RRC重建流程，若UE上下文不能被恢复但有资源，基站会触发哪条信令？",
        "options": ["RRC Setup", "不回任何消息", "RRC establishment", "RRC Reject"],
        "answer": 0,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "5G协议与信令", "answer_lookup": "RRC Setup"},
        [],
    )

    assert "UE上下文不能恢复但还有资源" in explanation
    assert "基站会转入重新建立连接流程" in explanation
    assert "所以选 A. RRC Setup" in explanation


def test_build_core_explanation_uses_reasoning_for_srb3_not_established() -> None:
    question = {
        "type": "single",
        "question": "1. 5G系统中，NSA场景下，当SRB3未建立时，SCG的测量结果，UE通过下面哪条消息上报给网络？",
        "options": ["UL Information Transfer MRDC", "RRC Reconfiguration Complete", "RRC Connection Reconfiguration Complete", "Measurement Report"],
        "answer": 0,
    }

    explanation = merge_all_banks.build_core_explanation(
        question,
        {"keypoint": "5G协议与信令", "answer_lookup": "UL Information Transfer MRDC"},
        [],
    )

    assert "SRB3未建立时" in explanation
    assert "SCG测量结果要通过UL Information Transfer MRDC上报" in explanation
    assert "所以选 A. UL Information Transfer MRDC" in explanation


def test_deduplicate_questions_removes_exact_duplicates_only() -> None:
    questions = [
        {
            "type": "single",
            "question": "重复题目",
            "options": ["A", "B", "C", "D"],
            "answer": 1,
            "explanation": "来源: repo1",
        },
        {
            "type": "single",
            "question": "  重复题目  ",
            "options": ["A", "B", "C", "D"],
            "answer": 1,
            "explanation": "来源: repo2",
        },
        {
            "type": "single",
            "question": "重复题目",
            "options": ["A", "B", "C", "D"],
            "answer": 2,
            "explanation": "来源: repo3",
        },
    ]

    deduped, removed = merge_all_banks.deduplicate_questions(questions)

    assert removed == 1
    assert len(deduped) == 2
    assert deduped[0]["explanation"] == "来源: repo1"
    assert deduped[1]["answer"] == 2


def test_main_deduplicates_exported_questions(tmp_path: Path) -> None:
    root = tmp_path / "DTCup_Prep"
    repo = root / "sample-repo"
    export_dir = root / "fusion" / "exports"
    repo.mkdir(parents=True)
    export_dir.mkdir(parents=True)

    duplicate_payload = {
        "questions": [
            {
                "question": "同一道题",
                "options": ["选项一", "选项二"],
                "answer": 0,
            },
            {
                "question": " 同一道题 ",
                "options": ["选项一", "选项二"],
                "answer": 0,
            },
            {
                "question": "同一道题",
                "options": ["选项一", "选项二"],
                "answer": 1,
            },
        ]
    }
    (repo / "dup.json").write_text(json.dumps(duplicate_payload, ensure_ascii=False), encoding="utf-8")

    merge_all_banks.ROOT = root
    merge_all_banks.EXPORT_DIR = export_dir
    merge_all_banks.REPOS = [repo]
    original_enable_dedup = merge_all_banks.ENABLE_DEDUP
    merge_all_banks.ENABLE_DEDUP = True
    try:
        merge_all_banks.main()
    finally:
        merge_all_banks.ENABLE_DEDUP = original_enable_dedup

    mega_bank = json.loads((export_dir / "mega_bank.json").read_text(encoding="utf-8"))
    report = (export_dir / "mega_bank_report.txt").read_text(encoding="utf-8")

    assert len(mega_bank["questions"]) == 2
    assert "REMOVED_DUPLICATES: 1" in report


def test_main_keeps_duplicates_when_dedup_disabled(tmp_path: Path) -> None:
    root = tmp_path / "DTCup_Prep"
    repo = root / "sample-repo"
    export_dir = root / "fusion" / "exports"
    repo.mkdir(parents=True)
    export_dir.mkdir(parents=True)

    duplicate_payload = {
        "questions": [
            {
                "question": "同一道题",
                "options": ["选项一", "选项二"],
                "answer": 0,
            },
            {
                "question": "同一道题",
                "options": ["选项一", "选项二"],
                "answer": 0,
            },
        ]
    }
    (repo / "dup.json").write_text(json.dumps(duplicate_payload, ensure_ascii=False), encoding="utf-8")

    merge_all_banks.ROOT = root
    merge_all_banks.EXPORT_DIR = export_dir
    merge_all_banks.REPOS = [repo]
    original_enable_dedup = merge_all_banks.ENABLE_DEDUP
    merge_all_banks.ENABLE_DEDUP = False
    try:
        merge_all_banks.main()
    finally:
        merge_all_banks.ENABLE_DEDUP = original_enable_dedup

    mega_bank = json.loads((export_dir / "mega_bank.json").read_text(encoding="utf-8"))
    report = (export_dir / "mega_bank_report.txt").read_text(encoding="utf-8")

    assert len(mega_bank["questions"]) == 2
    assert "REMOVED_DUPLICATES: 0" in report


def test_main_skips_synced_mega_bank_json_but_keeps_other_banks(tmp_path: Path) -> None:
    root = tmp_path / "DTCup_Prep"
    quiz_repo = root / "DTCUP-Quiz" / "src" / "data"
    export_dir = root / "fusion" / "exports"
    quiz_repo.mkdir(parents=True)
    export_dir.mkdir(parents=True)

    synced_mega_bank = {
        "name": "DTCUP Mega Bank",
        "questions": [
            {
                "question": "这道题不应被重新导入",
                "options": ["选项一", "选项二"],
                "answer": 0,
            }
        ],
    }
    raw_bank = {
        "name": "原始题库",
        "questions": [
            {
                "question": "这道题应被保留",
                "options": ["选项一", "选项二"],
                "answer": 1,
            }
        ],
    }
    (quiz_repo / "dtcup2025.json").write_text(json.dumps(synced_mega_bank, ensure_ascii=False), encoding="utf-8")
    (quiz_repo / "other_bank.json").write_text(json.dumps(raw_bank, ensure_ascii=False), encoding="utf-8")

    merge_all_banks.ROOT = root
    merge_all_banks.EXPORT_DIR = export_dir
    merge_all_banks.REPOS = [root / "DTCUP-Quiz"]

    merge_all_banks.main()

    mega_bank = json.loads((export_dir / "mega_bank.json").read_text(encoding="utf-8"))
    report = (export_dir / "mega_bank_report.txt").read_text(encoding="utf-8")

    assert [item["question"] for item in mega_bank["questions"]] == ["这道题应被保留"]
    assert "DTCUP-Quiz: 1" in report


def test_main_uses_only_curated_dtcup_2026_prep_sources(tmp_path: Path) -> None:
    root = tmp_path / "DTCup_Prep"
    repo = root / "dtcup-2026-prep"
    export_dir = root / "fusion" / "exports"
    bank_dir = repo / "真题题库"
    phase_dir = repo / "Phase1-基础入门"
    bank_dir.mkdir(parents=True)
    phase_dir.mkdir(parents=True)
    export_dir.mkdir(parents=True)

    (phase_dir / "Day1.md").write_text(
        "\n".join(
            [
                "1. 这道 Phase 笔记题不应被导入",
                "A. 选项一",
                "B. 选项二",
                "答案: A",
            ]
        ),
        encoding="utf-8",
    )
    (bank_dir / "真题-按知识点分类.md").write_text(
        "\n".join(
            [
                "# 分类题库",
                "",
                "**[来源：卷一]** 1. 真题主集题目",
                "",
                "- **A.** 选项一",
                "- **B.** 选项二 ✓",
                "【答案】B",
            ]
        ),
        encoding="utf-8",
    )
    (bank_dir / "真题-易错多选题.md").write_text("# 空文件\n", encoding="utf-8")
    (bank_dir / "真题-高频重复题.md").write_text("# 空文件\n", encoding="utf-8")

    merge_all_banks.ROOT = root
    merge_all_banks.EXPORT_DIR = export_dir
    merge_all_banks.REPOS = [repo]

    merge_all_banks.main()

    mega_bank = json.loads((export_dir / "mega_bank.json").read_text(encoding="utf-8"))
    report = (export_dir / "mega_bank_report.txt").read_text(encoding="utf-8")

    assert [item["question"] for item in mega_bank["questions"]] == ["1. 真题主集题目"]
    assert "dtcup-2026-prep: 1" in report
